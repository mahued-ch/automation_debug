import pyautogui
import time
import os
import cv2
import sys
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import win32com.client
import webbrowser
import win32clipboard
import io
import ctypes
import glob
import numpy as np
from Automation_Variables import configuracion 
import Automation_Variables

###--------------------------------------------------------------------------------
def limpiar_pantalla():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')   

# def busca_en_pantalla(imagenes, intervalo, reintentos, confidence=0.8):
#     """
#     Busca en pantalla cualquiera de las imágenes dadas, con intentos y pausas.
#     Si no encuentra nada, guarda captura de pantalla en carpeta 'logs' y registra log.

#     :param imagenes: Lista de rutas de imágenes a buscar.
#     :param intervalo: Tiempo en segundos entre reintentos.
#     :param reintentos: Número máximo de reintentos.
#     :param confidence: Nivel de confianza para el reconocimiento de imagen (0.0 a 1.0).
#     :return: Coordenadas (x, y) si encuentra alguna imagen, o None.
#     """
#     location = None
#     intentos = 0

#     while location is None and intentos < reintentos:
#         for imagen in imagenes:
#             try:
#                 location = pyautogui.locateCenterOnScreen(imagen, confidence=confidence)
#                 if location:
#                     break
#             except Exception:
#                 pass
#         if location is None:
#             intentos += 1
#             time.sleep(intervalo)

#     if location is None:
#         # Crear carpeta 'logs' si no existe
#         ruta_logs = os.path.join(os.path.dirname(__file__), "logs")
#         os.makedirs(ruta_logs, exist_ok=True)

#         # Nombre de archivo con nombres de imágenes y timestamp
#         nombres = "_".join([os.path.splitext(os.path.basename(img))[0] for img in imagenes])
#         timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#         captura_path = os.path.join(ruta_logs, f"{nombres}_{timestamp}.png")
#         log_path = os.path.join(ruta_logs, "busqueda.log")

#         # Guardar captura de pantalla
#         pyautogui.screenshot(captura_path)

#         # Escribir en log
#         with open(log_path, "a", encoding="utf-8") as f:
#             f.write(f"[{timestamp}] No se encontró ninguna de las imágenes: {imagenes}\n")
#             f.write(f"  Captura guardada en: {captura_path}\n\n")

#     return location

def busca_en_pantalla(imagenes, intervalo, reintentos, confidence=0.8):
    """
    Busca una lista de imágenes en la pantalla usando OpenCV.
    Guarda captura en 'logs/' si ninguna imagen es encontrada.

    :param imagenes: Lista de rutas a imágenes.
    :param intervalo: Tiempo entre reintentos (segundos).
    :param reintentos: Número máximo de intentos.
    :param confidence: Nivel mínimo de coincidencia (0.0 a 1.0).
    :return: Coordenadas (x, y) si encuentra alguna imagen, o None.
    """
    location = None
    intentos = 0

    while location is None and intentos < reintentos:
        screenshot = pyautogui.screenshot()
        screen_np = cv2.cvtColor(np.array(screenshot), cv2.COLOR_RGB2GRAY)

        for imagen_path in imagenes:
            template = cv2.imread(imagen_path, cv2.IMREAD_GRAYSCALE)
            if template is None:
                print(f"⚠️ No se pudo cargar la imagen: {imagen_path}")
                continue

            result = cv2.matchTemplate(screen_np, template, cv2.TM_CCOEFF_NORMED)
            _, max_val, _, max_loc = cv2.minMaxLoc(result)

            if max_val >= confidence:
                # Coordenadas centradas del template
                h, w = template.shape
                location = (max_loc[0] + w // 2, max_loc[1] + h // 2)
                break  # sale del for de imágenes

        if location is None:
            intentos += 1
            time.sleep(intervalo)

    if location is None:
        ruta_logs = os.path.join(os.path.dirname(__file__), "logs")
        os.makedirs(ruta_logs, exist_ok=True)

        nombres = "_".join([os.path.splitext(os.path.basename(img))[0] for img in imagenes])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        captura_path = os.path.join(ruta_logs, f"{nombres}_{timestamp}.png")
        log_path = os.path.join(ruta_logs, "busqueda_opencv.log")

        # Guardar captura y log
        pyautogui.screenshot(captura_path)
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] No se encontró ninguna de las imágenes: {imagenes}\n")
            f.write(f"  Captura guardada en: {captura_path}\n\n")

    return location

###--------------------------------------------------------------------------------
def buscar_centro_en_pantalla(lista_imagenes, confidence=0.8):
    """
    Busca el centro de la primera coincidencia encontrada entre varias imágenes.

    :param lista_imagenes: Lista de rutas de imágenes a buscar.
    :param confidence: Nivel de confianza para la búsqueda.
    :return: Coordenadas (x, y) del centro de la imagen encontrada, o None si no encuentra.
    """
    for imagen in lista_imagenes:
        try:
            location = pyautogui.locateCenterOnScreen(imagen, confidence=confidence)
            if location:
                return location
        except Exception:
            pass
    return None

###--------------------------------------------------------------------------------
def buscar_todas_ocurrencias(imagen, confidence=0.8):
#    Busca todas las posiciones de una imagen en la pantalla.
    posiciones = None
    while (posiciones == None) :
        try:
            posiciones = list(pyautogui.locateAllOnScreen(imagen, confidence=confidence))
        except Exception:
            pass
            break
    return posiciones

###--------------------------------------------------------------------------------

def convert_to_excel(input_file):
    output_file = input_file.rsplit('.', 1)[0] + '.xlsx'
    resultado = {"archivo_salida": output_file, "analisis": {}}

    try:
        df = pd.read_csv(input_file, sep='\t', encoding='utf-16', engine='python', on_bad_lines='skip', skiprows=1)

        # Eliminar columnas sin nombre
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')] 

        # Guardar en Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Convertir a tabla en Excel
        wb = load_workbook(output_file)
        ws = wb.active
        tabla = Table(displayName="TablaDatos", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        # Ajustar el ancho de las columnas automáticamente
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))  # Convertir a string antes de medir
                except:
                    pass
            adjusted_width = (max_length + 2)  # Añadir un margen adicional
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)

        # Análisis de la columna "ID mensaje"
        if 'ID mensaje' in df.columns:
            id_mensaje_counts = df['ID mensaje'].value_counts().to_dict()
            resultado["analisis"]["ID mensaje"] = id_mensaje_counts

        return resultado  # Devuelve un diccionario con el nombre del archivo y el análisis

    except Exception as e:
        return {"error": str(e)}

###--------------------------------------------------------------------------------

def enviar_email(resultado, fecha_inicio, fecha_fin, destinatario, archivo_adjunto, servidor_smtp, puerto, usuario, contraseña):
    # Formatear fechas en el asunto
    asunto = f"LOG DEBUG SAP S4P {fecha_inicio} al {fecha_fin}"

    # Construir cuerpo del mensaje
    tipos = resultado["analisis"]
    tipos_texto = "\n".join([f"{tipo}\t\t{cantidad} registros" for tipo, cantidad in tipos.items()])

    # Verificar si hay eventos relevantes
    tipos_relevantes = {"BU0", "BU1", "EU3"}
    tipos_encontrados = set(tipos.keys())
    
    if tipos_encontrados.issubset(tipos_relevantes):
        mensaje_final = "Por lo que no se encontró ningún tipo que sea relevante."
    else:
        mensaje_final = "Se encontraron eventos relevantes."

    cuerpo_mensaje = f"""Adjunto envío el LOG DEBUG SAP S4P correspondiente a los días {fecha_inicio} al {fecha_fin}.\n
Se encontraron los siguientes tipos:\n
{tipos_texto}\n
{mensaje_final}
"""

    # Configurar el mensaje
    msg = MIMEMultipart()
    msg["From"] = usuario
    msg["To"] = destinatario
    msg["Subject"] = asunto
    msg.attach(MIMEText(cuerpo_mensaje, "plain"))

    # Adjuntar archivo
    with open(archivo_adjunto, "rb") as adjunto:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(adjunto.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename={archivo_adjunto}")
        msg.attach(part)

    # Enviar correo
    try:
        with smtplib.SMTP(servidor_smtp, puerto) as server:
            server.starttls()
            server.login(usuario, contraseña)
            server.sendmail(usuario, destinatario, msg.as_string())
        print("Correo enviado correctamente.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")
        
###--------------------------------------------------------------------------------
        
def obtener_archivos_adicionales(directorio, prefijo, extensiones):
    """
    Busca archivos en un directorio que empiecen con un prefijo dado y tengan las extensiones especificadas.
    """
    archivos = []
    for archivo in os.listdir(directorio):
        if archivo.startswith(prefijo) and archivo.endswith(tuple(extensiones)):
            archivos.append(os.path.join(directorio, archivo))
    return archivos
       
###--------------------------------------------------------------------------------
        
def abrir_correo_outlook(configuracion, titulo, resultado, fecha_inicio, fecha_fin, destinatario, archivos_adjuntos, directorio, prefijo, extensiones):
    
    # primero armamos los textos antes de abrir outlook
    subject = f"{titulo} {fecha_inicio} al {fecha_fin}"
    
    # Extraer correctamente los datos de análisis
    analisis = resultado.get("analisis", {})

    # Si los datos están anidados bajo "ID mensaje", extraerlos
    if isinstance(analisis, dict) and "ID mensaje" in analisis:
        analisis = analisis["ID mensaje"]

    # Ordenar los tipos de mayor a menor cantidad
    analisis_ordenado = dict(
        sorted(((k.strip(), v) for k, v in analisis.items()), key=lambda x: x[0], reverse=False)
    )

    # Limpiar nombres de tipos y formatear correctamente
    tipos_texto = "\n".join([f"{tipo.strip()}\t\t{cantidad} registros" for tipo, cantidad in analisis_ordenado.items()])

#    tipos_relevantes = {"BU0", "BU1", "EU3", "BUS"}
    tipos_relevantes = set(configuracion["tipos"]["no_relevantes"])
    tipos_encontrados = {tipo.strip() for tipo in analisis.keys()}  # Limpiar los nombres antes de comparar

    if not analisis:  # Si el análisis está vacío
        mensaje_final = "No se encontró ningún evento en el periodo especificado."
    elif tipos_encontrados.issubset(tipos_relevantes):
        mensaje_final = "Por lo que no se encontró ningún tipo que sea relevante."
    else:
        mensaje_final = "Se encontraron eventos relevantes."

    cuerpo_mensaje = f"""Adjunto envío el {titulo} correspondiente a los días {fecha_inicio} al {fecha_fin}.\n
Se encontraron los siguientes tipos:\n
{tipos_texto}\n
{mensaje_final}
"""
    
    # Crear la instancia de Outlook
    outlook = None
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        print("Outlook se ha iniciado correctamente.")
    except Exception as e:
        print(f"Error al iniciar Outlook: {e}")

    if outlook: 
        mail = outlook.CreateItem(0)  # 0 = Correo

        # Asunto del correo
        mail.Subject = subject

        mail.Body = cuerpo_mensaje
        mail.To = destinatario  # Se puede dejar vacío si el usuario debe ingresarlo

        # Adjuntar los archivos proporcionados por el usuario
        for archivo in archivos_adjuntos:
            mail.Attachments.Add(archivo)

        # Buscar archivos adicionales en el directorio especificado y agregar los encontrados
        archivos_extra = obtener_archivos_adicionales(directorio, prefijo, ['.png', '.docx'])
        for archivo in archivos_extra:
            mail.Attachments.Add(archivo)

        # Mostrar el correo para que el usuario lo revise y lo envíe manualmente
        mail.Display()
    else:
        mailto_link = f"mailto:{destinatario}?subject={subject}&body={cuerpo_mensaje}"
        webbrowser.open(mailto_link)
        os.startfile(directorio)  

            
###--------------------------------------------------------------------------------

def convert_to_excel_fiori(input_file):
    output_file = input_file.rsplit('.', 1)[0] + '.xlsx'
    resultado = {"archivo_salida": output_file, "analisis": {}}

    try:
        df = pd.read_csv(input_file, sep='\t', encoding='utf-16', engine='python', on_bad_lines='skip', skiprows=9)

        # Eliminar columnas sin nombre
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')] 

        # Guardar en Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Convertir a tabla en Excel
        wb = load_workbook(output_file)
        ws = wb.active
        tabla = Table(displayName="TablaDatos", ref=ws.dimensions)
        estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)

        # Ajustar el ancho de las columnas automáticamente
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))  # Convertir a string antes de medir
                except:
                    pass
            adjusted_width = (max_length + 2)  # Añadir un margen adicional
            ws.column_dimensions[column].width = adjusted_width

        wb.save(output_file)

        # Análisis de la columna "Ár." y "N"
        if 'Ár.' in df.columns and 'N' in df.columns:
            df['ID mensaje'] = df['Ár.'].astype(str).str.strip() + df['N'].astype(str).str.strip()  # Concatenar con espacio
            concatenado_counts = df['ID mensaje'].value_counts().to_dict()  # Contar ocurrencias
            resultado["analisis"]["ID mensaje"] = concatenado_counts  # Guardar en el resultado

        return resultado  # Devuelve un diccionario con el nombre del archivo y el análisis

    except Exception as e:
        return {"error": str(e)}

###--------------------------------------------------------------------------------

def captura_pantalla(region):
    """
    Captura la pantalla en la región especificada.
    
    :param region: Tupla (x, y, ancho, alto) que define la zona a capturar.
    :return: Imagen capturada (PIL.Image).
    """
    region_porcentual = calcular_region_porcentual(*region)
    
        # Captura pantalla
    imagen = pyautogui.screenshot(region=region_porcentual)

 # Buscar el siguiente número disponible
    if Automation_Variables.running_test == 1:
        # Fecha y hora actual
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")        
        
        patron = f"output/{timestamp}_{Automation_Variables.file_output}_*.png"
        archivos_existentes = glob.glob(patron)
        # Calcular el siguiente consecutivo
        siguiente = len(archivos_existentes) + 1        
        
        # Nombre del archivo con relleno de ceros (ej. 001, 002)
#        nombre_archivo = f"output/screenshot_{siguiente:03}.png"
        nombre_archivo = f"output/{timestamp}_{Automation_Variables.file_output}_{siguiente:02}.png"
            
    # Guardar imagen
        imagen.save(nombre_archivo)

    return imagen

###--------------------------------------------------------------------------------
        
def espera_cambio_pantalla(intervalo, reintentos, region, captura_inicial):
    """
    Espera hasta que la imagen en una región específica de la pantalla cambie.
    Compara con la captura inicial tomada previamente.
    
    :param intervalo: Tiempo en segundos entre cada verificación.
    :param reintentos: Cantidad máxima de intentos.
    :param region: Coordenadas (x, y, ancho, alto) donde buscar la imagen.
    :param captura_inicial: Imagen tomada previamente para comparación.
    :param confidence: Nivel de confianza para la comparación de imágenes.
    :return: True si detecta un cambio, False si no detecta cambios tras los intentos.
    """
    conta = 0

    # Esperar y verificar cambios entre las capturas
    while conta < reintentos:
        time.sleep(intervalo)
        conta += 1
        
        # Capturar la nueva imagen de la región
        screenshot = captura_pantalla(region)
        
        # Comparar con la captura inicial
        if not imagenes_son_iguales(captura_inicial, screenshot):
            print(f"[{conta}] Cambio detectado en la pantalla.")
            copiar_imagen_al_clipboard(screenshot)
            return True  # Se detectó un cambio, salir del bucle
        
        print(f"[{conta}] La imagen sigue igual. Esperando...")

    # Si se alcanzaron los reintentos sin cambios, mostrar mensaje
    print("Se alcanzó el límite de reintentos y la pantalla NO cambió.")
    return False

###--------------------------------------------------------------------------------

def imagenes_son_iguales(img1, img2):
    """
    Compara dos imágenes y determina si son iguales.
    
    :param img1: Primera imagen (PIL.Image).
    :param img2: Segunda imagen (PIL.Image).
    :return: True si son iguales, False si son diferentes.
    """
    return list(img1.getdata()) == list(img2.getdata())

###--------------------------------------------------------------------------------
        
def calcular_region_porcentual(porcentaje_x, porcentaje_y, porcentaje_w, porcentaje_h):
    """
    Calcula la región de captura en base a porcentajes de la pantalla.
    
    :param porcentaje_x: Porcentaje de la pantalla donde comienza la región en X.
    :param porcentaje_y: Porcentaje de la pantalla donde comienza la región en Y.
    :param porcentaje_w: Porcentaje del ancho de la pantalla que ocupará la región.
    :param porcentaje_h: Porcentaje del alto de la pantalla que ocupará la región.
    :return: Región de captura ajustada a la resolución actual.
    """
    ancho_pantalla, alto_pantalla = obtener_resolucion()
    
    x = int(ancho_pantalla * porcentaje_x)
    y = int(alto_pantalla * porcentaje_y)
    w = int(ancho_pantalla * porcentaje_w)
    h = int(alto_pantalla * porcentaje_h)

    region_ajustada = ajustar_region_por_escala((x, y, w, h))

#    return (x, y, w, h)
    return region_ajustada

###--------------------------------------------------------------------------------
def obtener_escala_dpi():
    """
    Obtiene la escala de pantalla en Windows (Ej: 100%, 125%, 150%).
    :return: Factor de escala (Ej: 1.0 para 100%, 1.25 para 125%).
    """
    try:
        awareness = ctypes.c_int()
        ctypes.windll.shcore.SetProcessDpiAwareness(2)  # Habilitar DPI Awareness
        factor_escala = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100.0
        return factor_escala
    except AttributeError:
        return 1.0  # En caso de error, asumir 100%
    
###--------------------------------------------------------------------------------
def obtener_resolucion():
    ancho, alto = pyautogui.size()  # Obtiene la resolución actual de la pantalla
    return ancho, alto

###--------------------------------------------------------------------------------

def ajustar_region_por_escala(region):
    """
    Ajusta una región de captura según la escala de pantalla (DPI).
    
    :param region: Tupla (x, y, ancho, alto) de la región en coordenadas normales.
    :return: Región ajustada según el factor de escala.
    """
    escala = obtener_escala_dpi()
    x, y, w, h = region
    return int(x / escala), int(y / escala), int(w / escala), int(h / escala)

###--------------------------------------------------------------------------------
def copiar_imagen_al_clipboard(imagen):
    """
    Copia una imagen PIL al portapapeles para que pueda pegarse en Paint u otros programas.
    
    :param imagen: Objeto de imagen PIL.
    """
    # Crear un buffer de memoria para la imagen
    output = io.BytesIO()
    imagen.convert("RGB").save(output, format="BMP")  # Guardar en formato BMP
    data = output.getvalue()[14:]  # Quitar cabecera BMP
    output.close()
    
    # Abrir el portapapeles y copiar la imagen
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()
    print("✅ Imagen copiada al portapapeles. Puedes pegarla en Paint.")        
    
###--------------------------------------------------------------------------------
def escape_keyboard_input(text):
    # Escapa % primero
    text = text.replace('%', '%%')
    # Luego escapamos los otros caracteres que keyboard interpreta como comandos
    special_chars = {'{': '{{', '}': '}}', '+': '{+}', '^': '{^}', '~': '{~}'}
    for char, replacement in special_chars.items():
        text = text.replace(char, replacement)
    return text    